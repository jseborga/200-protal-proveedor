"""Exportacion de un presupuesto al Excel canonico de `ssa_construction_apu`.

El asistente `apu.import.wizard` del modulo Odoo lee un libro con 7 hojas y
resuelve las relaciones por TEXTO (columnas `*_ref`), no por id. Por eso lo
unico que importa es que las refs sean consistentes entre hojas; aqui se
generan refs cortas y legibles (`R1`, `I1`, `INS1`).

Convenciones que hay que respetar al pie de la letra:

* Las lineas de tipo `sub` NO producen una fila en INSUMOS: Odoo autogenera
  ese insumo desde la partida complementaria y lo indexa con la clave
  sintetica ``__sub__<nombre del item complementario>``. Se replica esa
  clave en `ITEM_LINEAS.insumo_ref` para que enlace.
* Dos lineas que apuntan al mismo insumo con el mismo precio son el MISMO
  recurso (una sola fila en INSUMOS). Si el precio difiere son recursos
  distintos: un presupuesto puede tener el mismo material a dos precios
  (dos proveedores, dos momentos) y colapsarlos falsearia el APU.
* Los numeros se escriben ya redondeados con la precision de la obra
  (`ApuProject.decimals_*`), no con la del float de Python: el Excel es el
  documento que viaja al ERP y debe cuadrar contra la planilla impresa.

La funcion no toca la base de datos: recibe el proyecto ya cargado con sus
relaciones (rubros, items, lines, computos).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .apu_engine import apu_round

# Prefijo con el que Odoo indexa el insumo autogenerado de una partida
# complementaria consumida via linea type='sub'.
SUB_REF_PREFIX = "__sub__"

# Tipos de recurso que SI viajan a la hoja INSUMOS ('sub' queda fuera).
EXPORTABLE_RESOURCE_TYPES = ("mat", "mo", "eq")

# Tipos validos de fila de plantilla (identicos a Odoo).
TEMPLATE_LINE_TYPES = ("sum_mat", "sum_mo", "sum_eq", "percent", "formula")

# Cabeceras exactas por hoja. El orden es contractual: el wizard de Odoo
# localiza las columnas por nombre, pero cualquier renombre rompe la carga.
SHEET_HEADERS: dict[str, tuple[str, ...]] = {
    "CONFIG": ("clave", "valor"),
    "PLANTILLA_CALC": (
        "seq", "codigo", "descripcion", "tipo", "valor_pct", "formula",
        "es_precio_final",
    ),
    "RUBROS": ("ref", "nombre", "secuencia"),
    "INSUMOS": ("ref", "nombre", "tipo", "unidad", "precio_unitario"),
    "ITEMS": (
        "ref", "nombre", "rubro_ref", "unidad", "precio_referencia",
        "es_complementario", "estado", "secuencia", "cantidad_contrato",
    ),
    "ITEM_LINEAS": ("item_ref", "insumo_ref", "rendimiento", "notas"),
    "COMPUTOS": ("item_ref", "sector", "piezas", "largo", "ancho", "alto"),
}

# Orden de creacion de las hojas en el libro.
SHEET_ORDER = (
    "CONFIG", "PLANTILLA_CALC", "RUBROS", "INSUMOS", "ITEMS", "ITEM_LINEAS",
    "COMPUTOS",
)

# Clave de CONFIG que nombra la plantilla de calculo a usar/crear en Odoo.
CONFIG_TEMPLATE_KEY = "plantilla_calc"

_HEADER_FONT = Font(bold=True)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")

# Ancho aproximado por columna, solo cosmetico.
_COLUMN_WIDTHS = {
    "clave": 24, "valor": 34, "codigo": 12, "descripcion": 34, "tipo": 12,
    "formula": 28, "es_precio_final": 16, "ref": 10, "nombre": 46,
    "secuencia": 11, "unidad": 10, "precio_unitario": 16, "rubro_ref": 12,
    "precio_referencia": 18, "es_complementario": 18, "estado": 12,
    "cantidad_contrato": 18, "item_ref": 12, "insumo_ref": 30,
    "rendimiento": 14, "notas": 34, "sector": 28, "piezas": 10, "largo": 10,
    "ancho": 10, "alto": 10, "valor_pct": 12, "seq": 8,
}


class ApuExportError(ValueError):
    """Datos del presupuesto incompatibles con el formato canonico."""


def _number_format(decimals: int) -> str:
    """Formato de celda Excel para N decimales ('0.00', '0.000'...)."""
    decimals = max(0, int(decimals or 0))
    return "0" if decimals == 0 else "0." + "0" * decimals


def _clean(value) -> str:
    """Texto limpio para una celda.

    None se convierte en cadena vacia (celda en blanco), nunca en el literal
    "None", que Odoo importaria como si fuera un nombre.
    """
    if value is None:
        return ""
    return str(value).strip()


def _decimals(project, attr: str, default: int) -> int:
    value = getattr(project, attr, None)
    if value is None:
        return default
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return default


def _bool_cell(value) -> int:
    """Booleano como 1/0: es lo que el wizard de Odoo sabe interpretar."""
    return 1 if value else 0


def _sorted_by_sequence(records) -> list:
    """Ordena por (sequence, id) sin depender del ORDER BY del ORM."""
    records = list(records or [])
    return sorted(
        records,
        key=lambda r: (
            getattr(r, "sequence", 0) or 0,
            getattr(r, "id", 0) or 0,
        ),
    )


def _sub_ref(name: str) -> str:
    """Clave sintetica con la que Odoo indexa un sub-APU."""
    return f"{SUB_REF_PREFIX}{_clean(name)}"


class _ResourceIndex:
    """Asigna un `INSUMOS.ref` estable a cada recurso distinto.

    Dos lineas son el mismo recurso si coinciden tipo, insumo (o nombre, si
    la linea es manual), unidad y precio ya redondeado. El precio entra en la
    clave a proposito: el mismo material a dos precios son dos filas.
    """

    def __init__(self, decimals_price: int):
        self._decimals_price = decimals_price
        self._refs: dict[tuple, str] = {}
        self.rows: list[list] = []

    def ref_for(self, line) -> str:
        price = apu_round(getattr(line, "price_unit", 0.0), self._decimals_price)
        name = _clean(getattr(line, "name", "")) or "SIN NOMBRE"
        uom = _clean(getattr(line, "uom", "")) or "u"
        ltype = _clean(getattr(line, "type", ""))
        insumo_id = getattr(line, "insumo_id", None)

        identity = ("insumo", insumo_id) if insumo_id else ("manual", name.lower())
        key = (ltype, identity, uom.lower(), price)

        ref = self._refs.get(key)
        if ref is None:
            ref = f"INS{len(self._refs) + 1}"
            self._refs[key] = ref
            self.rows.append([ref, name, ltype, uom, price])
        return ref


def _write_sheet(wb: Workbook, name: str, rows: list[list], formats: dict[int, str]):
    """Crea una hoja con su cabecera y sus filas.

    `formats` mapea indice de columna (0-based) -> formato numerico Excel.
    """
    ws = wb.create_sheet(title=name)
    headers = SHEET_HEADERS[name]
    ws.append(list(headers))

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _COLUMN_WIDTHS.get(header, 16)

    for row in rows:
        ws.append(row)

    for col_idx, fmt in formats.items():
        letter = get_column_letter(col_idx + 1)
        for cell in ws[letter][1:]:  # se salta la cabecera
            cell.number_format = fmt

    ws.freeze_panes = "A2"
    return ws


def build_workbook(project, template=None) -> bytes:
    """Serializa un presupuesto al .xlsx que importa `apu.import.wizard`.

    Args:
        project: `ApuProject` YA cargado con `rubros`, `items`, y de cada item
            sus `lines` y `computos`. No se hacen consultas aqui.
        template: `ApuTemplate` con sus `lines` (opcional). Si es None se usa
            `project.template` y, si tampoco existe, PLANTILLA_CALC queda solo
            con la cabecera.

    Returns:
        Los bytes del libro Excel.

    Raises:
        ApuExportError: si una linea o una fila de plantilla trae un tipo que
            Odoo no sabe leer (fallar aqui es mejor que importar basura).
    """
    if project is None:
        raise ApuExportError("No hay presupuesto que exportar")

    if template is None:
        template = getattr(project, "template", None)

    d_price = _decimals(project, "decimals_price", 2)
    d_perf = _decimals(project, "decimals_performance", 3)
    d_total = _decimals(project, "decimals_total", 2)
    d_qty = _decimals(project, "decimals_qty", 4)

    # ── CONFIG ──────────────────────────────────────────────────
    config_rows = [
        ["decimals_price", d_price],
        ["decimals_performance", d_perf],
        ["decimals_subtotal", _decimals(project, "decimals_subtotal", 2)],
        ["decimals_total", d_total],
        ["decimals_qty", d_qty],
        [CONFIG_TEMPLATE_KEY, _clean(getattr(template, "name", ""))],
    ]

    # ── PLANTILLA_CALC ──────────────────────────────────────────
    plantilla_rows = []
    for tline in _sorted_by_sequence(getattr(template, "lines", []) if template else []):
        ttype = _clean(getattr(tline, "type", ""))
        if ttype not in TEMPLATE_LINE_TYPES:
            raise ApuExportError(
                f"Tipo de fila de plantilla no soportado por Odoo: {ttype!r}"
            )
        plantilla_rows.append([
            getattr(tline, "sequence", 0) or 0,
            _clean(getattr(tline, "code", "")),
            _clean(getattr(tline, "name", "")),
            ttype,
            float(getattr(tline, "value", 0.0) or 0.0),
            _clean(getattr(tline, "formula", "")),
            _bool_cell(getattr(tline, "is_total", False)),
        ])

    # ── RUBROS ──────────────────────────────────────────────────
    rubros = _sorted_by_sequence(getattr(project, "rubros", []))
    rubro_refs: dict[int, str] = {}
    rubro_rows = []
    for idx, rubro in enumerate(rubros, start=1):
        ref = f"R{idx}"
        rubro_refs[getattr(rubro, "id", None)] = ref
        rubro_rows.append([
            ref,
            _clean(getattr(rubro, "name", "")),
            getattr(rubro, "sequence", 0) or 0,
        ])

    # ── ITEMS / ITEM_LINEAS / COMPUTOS / INSUMOS ────────────────
    items = _sorted_by_sequence(getattr(project, "items", []))
    item_refs: dict[int, str] = {}
    item_names: dict[int, str] = {}
    for idx, item in enumerate(items, start=1):
        item_refs[getattr(item, "id", None)] = f"I{idx}"
        item_names[getattr(item, "id", None)] = _clean(getattr(item, "name", ""))

    resources = _ResourceIndex(d_price)
    item_rows: list[list] = []
    line_rows: list[list] = []
    computo_rows: list[list] = []

    for item in items:
        item_ref = item_refs[getattr(item, "id", None)]
        lines = _sorted_by_sequence(getattr(item, "lines", []))

        reference_price = getattr(item, "reference_price", None)
        if reference_price is None:
            # Sin composicion el APU no tiene de donde salir: se manda el
            # precio ya calculado para no perder el numero al migrar.
            reference_price = getattr(item, "unit_price", 0.0) if not lines else 0.0

        item_rows.append([
            item_ref,
            _clean(getattr(item, "name", "")),
            rubro_refs.get(getattr(item, "rubro_id", None), ""),
            _clean(getattr(item, "uom", "")) or "u",
            apu_round(reference_price, d_price),
            _bool_cell(getattr(item, "is_complementary", False)),
            _clean(getattr(item, "state", "")) or "draft",
            getattr(item, "sequence", 0) or 0,
            apu_round(getattr(item, "quantity", 0.0), d_qty),
        ])

        for line in lines:
            ltype = _clean(getattr(line, "type", ""))
            if ltype == "sub":
                linked_id = getattr(line, "linked_item_id", None)
                # Odoo indexa el sub-APU por el NOMBRE de la partida
                # complementaria. Si el enlace se perdio se usa el nombre
                # copiado en la linea, que es lo que vio el usuario.
                linked_name = item_names.get(linked_id) or _clean(
                    getattr(line, "name", "")
                )
                if not linked_name:
                    raise ApuExportError(
                        f"Linea sub-APU sin nombre en la partida {item_ref}"
                    )
                insumo_ref = _sub_ref(linked_name)
            elif ltype in EXPORTABLE_RESOURCE_TYPES:
                insumo_ref = resources.ref_for(line)
            else:
                raise ApuExportError(
                    f"Tipo de recurso no soportado por Odoo: {ltype!r}"
                )

            line_rows.append([
                item_ref,
                insumo_ref,
                apu_round(getattr(line, "quantity", 0.0), d_perf),
                _clean(getattr(line, "notes", "")),
            ])

        for computo in _sorted_by_sequence(getattr(item, "computos", [])):
            computo_rows.append([
                item_ref,
                _clean(getattr(computo, "name", "")),
                apu_round(getattr(computo, "pieces", 0.0), d_qty),
                apu_round(getattr(computo, "length", 0.0), d_qty),
                apu_round(getattr(computo, "width", 0.0), d_qty),
                apu_round(getattr(computo, "height", 0.0), d_qty),
            ])

    # ── Ensamblado del libro ────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # openpyxl crea una hoja vacia por defecto

    price_fmt = _number_format(d_price)
    perf_fmt = _number_format(d_perf)
    qty_fmt = _number_format(d_qty)

    payload = {
        "CONFIG": (config_rows, {}),
        "PLANTILLA_CALC": (plantilla_rows, {4: _number_format(4)}),
        "RUBROS": (rubro_rows, {}),
        "INSUMOS": (resources.rows, {4: price_fmt}),
        "ITEMS": (item_rows, {4: price_fmt, 8: qty_fmt}),
        "ITEM_LINEAS": (line_rows, {2: perf_fmt}),
        "COMPUTOS": (
            computo_rows, {2: qty_fmt, 3: qty_fmt, 4: qty_fmt, 5: qty_fmt},
        ),
    }
    for name in SHEET_ORDER:
        rows, formats = payload[name]
        _write_sheet(wb, name, rows, formats)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def default_filename(project) -> str:
    """Nombre de archivo sugerido: `APU_<codigo o nombre>.xlsx`."""
    raw = _clean(getattr(project, "code", "")) or _clean(
        getattr(project, "name", "")
    ) or "presupuesto"
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in raw)
    return f"APU_{safe[:60]}.xlsx"
