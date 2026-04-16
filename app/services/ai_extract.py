"""Extraccion de datos de cotizacion desde Excel, PDF o foto usando IA.

Soporta multiples proveedores de IA configurables desde admin y por empresa:
- Google AI Studio (gratis)
- OpenRouter (multi-modelo)
- Anthropic (Claude)
- OpenAI (GPT)

Resolucion de config: empresa → sistema (admin) → .env fallback
"""

import base64
import io
import json
import re

import httpx

from app.core.config import settings
from app.core.ai_providers import AI_PROVIDERS, DEFAULT_PROVIDER, get_provider_info


# ── Config resolution ──────────────────────────────────────────

async def resolve_ai_config(company_id: int | None = None) -> dict | None:
    """Resuelve la config de IA a usar: empresa → sistema → .env fallback.

    Returns dict con keys: provider, api_key, model, base_url, api_format
    o None si no hay config disponible.
    """
    from app.core.database import async_session
    from app.models.system_setting import SystemSetting

    # 1. Company config (si tiene company_id)
    if company_id:
        from app.models.company import Company
        async with async_session() as db:
            company = await db.get(Company, company_id)
            if company and company.extra_data:
                ai = company.extra_data.get("ai_config")
                if ai and ai.get("api_key"):
                    return _build_config(ai)

    # 2. System config (admin panel)
    async with async_session() as db:
        setting = await db.get(SystemSetting, "ai_config")
        if setting and setting.value and setting.value.get("api_key"):
            return _build_config(setting.value)

    # 3. Fallback to .env
    if settings.ai_api_key:
        return _build_config({
            "provider": settings.ai_provider,
            "api_key": settings.ai_api_key,
            "model": settings.ai_model,
        })

    return None


def _build_config(raw: dict) -> dict:
    """Normaliza la config raw a un dict consistente con provider info."""
    provider_key = raw.get("provider", DEFAULT_PROVIDER)
    provider = get_provider_info(provider_key) or get_provider_info(DEFAULT_PROVIDER)

    return {
        "provider": provider_key,
        "api_key": raw.get("api_key", ""),
        "model": raw.get("model") or provider["default_model"],
        "base_url": provider["base_url"],
        "api_format": provider["api_format"],
    }


# ── Excel extraction (no AI needed) ────────────────────────────
async def extract_quotation_data(
    content: bytes, filename: str, source: str,
    company_id: int | None = None,
) -> dict | None:
    """Extract quotation lines from uploaded file.

    For PDFs: tries text extraction first (pdfplumber), falls back to AI
    if the PDF is scanned/image-based or text parsing yields no items.
    """
    if source == "excel":
        return _extract_from_excel(content, filename)
    elif source == "pdf":
        # Try text-based extraction first (free, fast, reliable for digital PDFs)
        text_result = _extract_from_pdf_text(content, filename)
        if text_result and text_result.get("lines"):
            print(f"[AI] PDF text extraction succeeded: {len(text_result['lines'])} items")
            return text_result
        # Fallback: scanned PDF or text parsing failed → use AI vision
        print("[AI] PDF text extraction yielded no items, falling back to AI vision")
        return await _extract_with_ai(content, filename, source, company_id)
    elif source == "photo":
        return await _extract_with_ai(content, filename, source, company_id)
    return None


def _extract_from_excel(content: bytes, filename: str) -> dict | None:
    """Parse Excel file looking for price-list columns."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        return None

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None

    # Detect columns by header heuristics
    header = [str(c).lower().strip() if c else "" for c in rows[0]]
    col_map = _detect_columns(header)

    if "name" not in col_map:
        # Try second row as header
        header = [str(c).lower().strip() if c else "" for c in rows[1]]
        col_map = _detect_columns(header)
        rows = rows[1:]

    if "name" not in col_map:
        return None

    lines = []
    for row in rows[1:]:
        name_val = row[col_map["name"]] if col_map.get("name") is not None else None
        if not name_val or str(name_val).strip() == "":
            continue

        line = {"name": str(name_val).strip()}
        if "code" in col_map and col_map["code"] is not None:
            val = row[col_map["code"]]
            if val:
                line["code"] = str(val).strip()
        if "uom" in col_map and col_map["uom"] is not None:
            val = row[col_map["uom"]]
            if val:
                line["uom"] = str(val).strip()
        if "price" in col_map and col_map["price"] is not None:
            val = row[col_map["price"]]
            try:
                line["price"] = float(val)
            except (TypeError, ValueError):
                line["price"] = 0
        if "brand" in col_map and col_map["brand"] is not None:
            val = row[col_map["brand"]]
            if val:
                line["brand"] = str(val).strip()

        lines.append(line)

    wb.close()
    return {"lines": lines, "metadata": {"source": "excel", "filename": filename, "rows": len(lines)}}


def _detect_columns(header: list[str]) -> dict:
    """Heuristic column detection for price list spreadsheets."""
    col_map = {}
    name_keywords = ["descripcion", "nombre", "producto", "material", "insumo", "item", "detalle", "name"]
    code_keywords = ["codigo", "code", "cod", "ref", "sku"]
    uom_keywords = ["unidad", "uom", "und", "medida", "unit"]
    price_keywords = ["precio", "price", "unitario", "p.u.", "pu", "costo", "valor"]
    brand_keywords = ["marca", "brand", "fabricante"]

    for i, col in enumerate(header):
        if any(kw in col for kw in name_keywords) and "name" not in col_map:
            col_map["name"] = i
        elif any(kw in col for kw in code_keywords) and "code" not in col_map:
            col_map["code"] = i
        elif any(kw in col for kw in uom_keywords) and "uom" not in col_map:
            col_map["uom"] = i
        elif any(kw in col for kw in price_keywords) and "price" not in col_map:
            col_map["price"] = i
        elif any(kw in col for kw in brand_keywords) and "brand" not in col_map:
            col_map["brand"] = i

    return col_map


# ── PDF text extraction (no AI) ─────────────────────────────────

def _extract_from_pdf_text(content: bytes, filename: str) -> dict | None:
    """Extract data from a text-based PDF using pdfplumber.

    Works for digital PDFs (invoices, price lists generated by systems).
    Returns None for scanned/image-only PDFs.
    """
    try:
        import pdfplumber
    except ImportError:
        print("[PDF] pdfplumber not installed, skipping text extraction")
        return None

    try:
        pdf = pdfplumber.open(io.BytesIO(content))
    except Exception as e:
        print(f"[PDF] Failed to open: {e}")
        return None

    # Extract all text
    all_text = ""
    all_tables = []
    for page in pdf.pages:
        page_text = page.extract_text() or ""
        all_text += page_text + "\n"
        tables = page.extract_tables() or []
        all_tables.extend(tables)
    pdf.close()

    # If very little text, it's likely a scanned PDF → let AI handle it
    clean_text = all_text.strip()
    if len(clean_text) < 50:
        print(f"[PDF] Only {len(clean_text)} chars extracted, likely scanned")
        return None

    # Try table-based extraction first (most reliable for invoices)
    if all_tables:
        result = _parse_pdf_tables(all_tables, filename)
        if result and result.get("lines"):
            # Also try to extract supplier/doc info from text
            header_info = _parse_pdf_header(clean_text)
            if header_info.get("supplier"):
                result["supplier"] = header_info["supplier"]
            if header_info.get("document"):
                result["document"] = header_info["document"]
            return result

    # Fallback: line-by-line parsing of text
    result = _parse_pdf_lines(clean_text, filename)
    if result and result.get("lines"):
        header_info = _parse_pdf_header(clean_text)
        if header_info.get("supplier"):
            result["supplier"] = header_info["supplier"]
        if header_info.get("document"):
            result["document"] = header_info["document"]
        return result

    return None


def _parse_pdf_tables(tables: list, filename: str) -> dict | None:
    """Parse extracted PDF tables looking for price list data."""
    lines = []

    for table in tables:
        if not table or len(table) < 2:
            continue

        # First row as header
        header = [str(c).lower().strip() if c else "" for c in table[0]]
        col_map = _detect_columns(header)

        # If no name column found, try second row as header
        if "name" not in col_map and len(table) > 2:
            header = [str(c).lower().strip() if c else "" for c in table[1]]
            col_map = _detect_columns(header)
            data_rows = table[2:]
        else:
            data_rows = table[1:]

        if "name" not in col_map:
            continue

        for row in data_rows:
            if not row or len(row) <= col_map.get("name", 0):
                continue

            name_val = row[col_map["name"]]
            if not name_val or not str(name_val).strip():
                continue
            # Skip rows that look like subtotals
            name_str = str(name_val).strip()
            if any(kw in name_str.lower() for kw in ["total", "subtotal", "descuento", "iva", "neto"]):
                continue

            line = {"name": name_str}
            if "code" in col_map and col_map["code"] < len(row) and row[col_map["code"]]:
                line["code"] = str(row[col_map["code"]]).strip()
            if "uom" in col_map and col_map["uom"] < len(row) and row[col_map["uom"]]:
                line["uom"] = str(row[col_map["uom"]]).strip()
            if "price" in col_map and col_map["price"] < len(row):
                try:
                    price_str = str(row[col_map["price"]]).replace(",", "").replace("Bs", "").strip()
                    line["price"] = float(price_str)
                except (ValueError, TypeError):
                    line["price"] = 0
            if "brand" in col_map and col_map["brand"] < len(row) and row[col_map["brand"]]:
                line["brand"] = str(row[col_map["brand"]]).strip()

            lines.append(line)

    if not lines:
        return None

    return {
        "lines": lines,
        "metadata": {"source": "pdf_text", "filename": filename, "method": "table", "items_extracted": len(lines)},
    }


def _parse_pdf_lines(text: str, filename: str) -> dict | None:
    """Parse PDF text line by line looking for price patterns.

    Looks for patterns like:
    - "Cemento Portland IP-30   bls   58.50"
    - "10  Cemento Portland   bls  58.50  585.00"
    - "Fierro 12mm   kg   12.50"
    """
    lines_out = []
    # Pattern: optional quantity, name, optional uom, price
    # Common in Bolivian invoices/cotizaciones
    price_pattern = re.compile(
        r"^"
        r"(?:(\d+)\s+)?"  # optional leading quantity
        r"(.{5,80}?)"  # name (5-80 chars, non-greedy)
        r"\s+"
        r"(?:(m3|m2|ml|kg|tn|pza|bls|lt|gl|und|rollo|pliego|metro|unidad|bolsa|pieza|varilla|tubo|par|caja|saco|barra)\s+)?"  # optional uom
        r"(\d[\d.,]*\d|\d)"  # price (number with possible decimals)
        r"(?:\s+\d[\d.,]*\d?)?"  # optional total column
        r"\s*$",
        re.IGNORECASE | re.MULTILINE,
    )

    for line in text.split("\n"):
        line = line.strip()
        if not line or len(line) < 8:
            continue
        # Skip header/footer lines
        if any(kw in line.lower() for kw in [
            "total", "subtotal", "descuento", "iva", "neto", "factura",
            "fecha", "nit", "telefono", "direccion", "pagina", "page",
        ]):
            continue

        m = price_pattern.match(line)
        if m:
            qty_str, name, uom, price_str = m.groups()
            name = name.strip().rstrip("-–—.,:;")
            if len(name) < 3:
                continue
            try:
                price = float(price_str.replace(",", ""))
            except ValueError:
                continue
            if price <= 0 or price > 1_000_000:
                continue

            item = {"name": name, "price": price}
            if uom:
                item["uom"] = uom.lower()
            if qty_str:
                item["quantity"] = int(qty_str)
            lines_out.append(item)

    if len(lines_out) < 2:
        return None

    return {
        "lines": lines_out,
        "metadata": {"source": "pdf_text", "filename": filename, "method": "lines", "items_extracted": len(lines_out)},
    }


def _parse_pdf_header(text: str) -> dict:
    """Extract supplier and document info from PDF header text."""
    info: dict = {}

    # Try to find NIT
    nit_match = re.search(r"NIT[:\s]*(\d[\d.-]+\d)", text, re.IGNORECASE)
    nit = nit_match.group(1) if nit_match else None

    # Try to find document number
    doc_match = re.search(
        r"(?:factura|nota|proforma|cotizacion|recibo)[^:\n]*?[:\s#No.]*\s*([A-Z0-9][\w-]{2,20})",
        text, re.IGNORECASE,
    )
    doc_number = doc_match.group(1) if doc_match else None

    # Try to find date
    date_match = re.search(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", text)
    doc_date = None
    if date_match:
        d, m, y = date_match.groups()
        if len(y) == 2:
            y = "20" + y
        doc_date = f"{y}-{m.zfill(2)}-{d.zfill(2)}"

    # Detect document type
    text_lower = text[:500].lower()
    doc_type = None
    if "factura" in text_lower:
        doc_type = "factura"
    elif "proforma" in text_lower:
        doc_type = "proforma"
    elif "cotizacion" in text_lower or "cotización" in text_lower:
        doc_type = "cotizacion"
    elif "nota de venta" in text_lower or "nota de remision" in text_lower:
        doc_type = "nota_venta"
    elif "lista de precio" in text_lower:
        doc_type = "lista_precios"

    if doc_type or doc_number or doc_date:
        info["document"] = {"type": doc_type, "number": doc_number, "date": doc_date}

    # Try to find supplier name (usually first prominent line or after company-related keywords)
    # Look in first 5 lines for a business name
    first_lines = [l.strip() for l in text.split("\n")[:8] if l.strip() and len(l.strip()) > 3]
    supplier_name = None
    for line in first_lines:
        # Skip lines that are clearly not business names
        if any(kw in line.lower() for kw in [
            "factura", "fecha", "nit:", "tel:", "dir:", "pagina",
            "cotizacion", "nota de", "proforma", "recibo",
        ]):
            continue
        # A line that looks like a business name (not too long, not a number)
        if 4 < len(line) < 60 and not line.replace(".", "").replace(",", "").isdigit():
            supplier_name = line
            break

    if supplier_name or nit:
        info["supplier"] = {
            "name": supplier_name,
            "nit": nit,
            "phone": None,
            "address": None,
        }

    return info


# ── AI extraction (PDF / Photo) ─────────────────────────────────
EXTRACTION_PROMPT = """Extrae los items de este documento de construccion (cotizacion, factura, lista de precios, proforma, nota de venta, etc.).

Devuelve SOLO un JSON object con esta estructura:
{
  "supplier": {"name": "nombre del proveedor/empresa", "nit": "NIT si aparece", "phone": "telefono si aparece", "address": "direccion si aparece"},
  "document": {"type": "factura|cotizacion|proforma|lista_precios|nota_venta", "number": "numero de documento si existe", "date": "fecha si aparece (YYYY-MM-DD)"},
  "items": [
    {"name": "nombre del producto/material", "code": "codigo si existe", "uom": "unidad de medida (m3, kg, m2, pza, ml, bls, lt, gl, etc.)", "price": 0.00, "quantity": 1, "brand": "marca si se menciona"}
  ]
}

Reglas:
- "price" es el precio UNITARIO (si solo hay total y cantidad, divide total/cantidad)
- Si no identificas proveedor, pon supplier como null
- Si no identificas tipo de documento, pon document como null
- Los items siempre deben ser un array (vacio [] si no hay items)
- Limpia los nombres: sin codigos internos, sin caracteres raros, primera letra mayuscula
- uom: normaliza a abreviaturas estandar (bls, kg, m3, m2, pza, ml, lt, gl, m, und, rollo, plza)

Ejemplo:
{"supplier": {"name": "Ferreteria El Constructor", "nit": "12345678", "phone": null, "address": "Av. Blanco Galindo 1234"}, "document": {"type": "factura", "number": "F-001234", "date": "2025-03-15"}, "items": [{"name": "Cemento Portland IP-30", "code": null, "uom": "bls", "price": 58.5, "quantity": 10, "brand": "Viacha"}]}

Si no puedes extraer datos, devuelve {"supplier": null, "document": null, "items": []}.
No incluyas texto adicional, SOLO el JSON."""


async def resolve_all_ai_configs(company_id: int | None = None) -> list[dict]:
    """Return all available AI configs in priority order for fallback."""
    from app.core.database import async_session
    from app.models.system_setting import SystemSetting

    configs = []
    seen_keys = set()

    # 1. Company config
    if company_id:
        from app.models.company import Company
        async with async_session() as db:
            company = await db.get(Company, company_id)
            if company and company.extra_data:
                ai = company.extra_data.get("ai_config")
                if ai and ai.get("api_key"):
                    c = _build_config(ai)
                    key = (c["provider"], c["api_key"][:8])
                    if key not in seen_keys:
                        configs.append(c)
                        seen_keys.add(key)

    # 2. System config (admin panel)
    async with async_session() as db:
        setting = await db.get(SystemSetting, "ai_config")
        if setting and setting.value and setting.value.get("api_key"):
            c = _build_config(setting.value)
            key = (c["provider"], c["api_key"][:8])
            if key not in seen_keys:
                configs.append(c)
                seen_keys.add(key)

    # 3. .env fallback (primary)
    if settings.ai_api_key:
        c = _build_config({
            "provider": settings.ai_provider,
            "api_key": settings.ai_api_key,
            "model": settings.ai_model,
        })
        key = (c["provider"], c["api_key"][:8])
        if key not in seen_keys:
            configs.append(c)
            seen_keys.add(key)

    # 4. Anthropic fallback — if ANTHROPIC_API_KEY is set, always add as last resort
    if settings.anthropic_api_key:
        c = _build_config({
            "provider": "anthropic",
            "api_key": settings.anthropic_api_key,
            "model": "claude-sonnet-4-20250514",
        })
        key = (c["provider"], c["api_key"][:8])
        if key not in seen_keys:
            configs.append(c)
            seen_keys.add(key)

    return configs


def _call_provider(
    content: bytes, filename: str, source: str, config: dict,
):
    """Route to the correct provider-specific function."""
    fmt = config["api_format"]
    if fmt == "google":
        return _call_google(content, filename, source, config)
    elif fmt == "openrouter":
        return _call_openrouter(content, filename, source, config)
    elif fmt == "anthropic":
        return _call_anthropic(content, filename, source, config)
    elif fmt == "openai":
        return _call_openai(content, filename, source, config)
    return None


async def _extract_with_ai(
    content: bytes, filename: str, source: str,
    company_id: int | None = None,
) -> dict | None:
    """Use AI to extract quotation data from PDF or image.

    Tries each configured provider in order (fallback on failure).
    """
    configs = await resolve_all_ai_configs(company_id)
    if not configs:
        return None

    last_error = None
    for config in configs:
        try:
            result = await _call_provider(content, filename, source, config)
            if result:
                return result
            print(f"[AI] {config['provider']}/{config['model']} returned no data, trying next...")
        except Exception as e:
            last_error = e
            print(f"[AI] {config['provider']}/{config['model']} failed: {e}, trying next...")

    if last_error:
        print(f"[AI] All providers failed. Last error: {last_error}")
    return None


async def _call_openrouter(
    content: bytes, filename: str, source: str, config: dict,
) -> dict | None:
    """Call OpenRouter API."""
    b64 = base64.b64encode(content).decode()
    media_type = "application/pdf" if source == "pdf" else "image/jpeg"

    messages = [
        {
            "role": "user",
            "content": [
                {"type": "text", "text": EXTRACTION_PROMPT},
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{media_type};base64,{b64}"},
                },
            ],
        }
    ]

    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            config["base_url"],
            headers={
                "Authorization": f"Bearer {config['api_key']}",
                "HTTP-Referer": "https://apu-marketplace.com",
                "Content-Type": "application/json",
            },
            json={"model": config["model"], "messages": messages, "max_tokens": 4096},
        )

    if resp.status_code != 200:
        raise Exception(f"OpenRouter {resp.status_code}: {resp.text[:200]}")

    data = resp.json()
    text_content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
    return _parse_ai_response(text_content, filename, source, config)


async def _call_anthropic(
    content: bytes, filename: str, source: str, config: dict,
) -> dict | None:
    """Call Anthropic Messages API directly."""
    b64 = base64.b64encode(content).decode()

    if source == "pdf":
        # Anthropic uses 'document' type for PDFs
        media_block = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
        }
    else:
        media_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": "image/jpeg", "data": b64},
        }

    messages = [
        {
            "role": "user",
            "content": [media_block, {"type": "text", "text": EXTRACTION_PROMPT}],
        }
    ]

    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            config["base_url"],
            headers={
                "x-api-key": config["api_key"],
                "anthropic-version": "2023-06-01",
                "Content-Type": "application/json",
            },
            json={"model": config["model"], "messages": messages, "max_tokens": 4096},
        )

    if resp.status_code != 200:
        raise Exception(f"Anthropic {resp.status_code}: {resp.text[:200]}")

    data = resp.json()
    text_content = ""
    for block in data.get("content", []):
        if block.get("type") == "text":
            text_content += block.get("text", "")

    return _parse_ai_response(text_content, filename, source, config)


async def _call_google(
    content: bytes, filename: str, source: str, config: dict,
) -> dict | None:
    """Google AI Studio — native Gemini generateContent API with ?key= auth."""
    b64 = base64.b64encode(content).decode()
    media_type = "application/pdf" if source == "pdf" else "image/jpeg"

    parts = [
        {"text": EXTRACTION_PROMPT},
        {"inline_data": {"mime_type": media_type, "data": b64}},
    ]

    model = config["model"]
    base = config["base_url"].rstrip("/")
    endpoint = f"{base}/models/{model}:generateContent?key={config['api_key']}"

    async with httpx.AsyncClient(timeout=90) as client:
        resp = await client.post(
            endpoint,
            headers={"Content-Type": "application/json"},
            json={
                "contents": [{"parts": parts}],
                "generationConfig": {"maxOutputTokens": 4096},
            },
        )

    if resp.status_code != 200:
        raise Exception(f"Google AI {resp.status_code}: {resp.text[:200]}")

    data = resp.json()
    text_content = ""
    for candidate in data.get("candidates", []):
        for part in candidate.get("content", {}).get("parts", []):
            if "text" in part:
                text_content += part["text"]

    return _parse_ai_response(text_content, filename, source, config)


async def _call_openai(
    content: bytes, filename: str, source: str, config: dict,
) -> dict | None:
    """OpenAI — standard OpenAI API with Bearer auth."""
    b64 = base64.b64encode(content).decode()
    media_type = "application/pdf" if source == "pdf" else "image/jpeg"

    messages = [
        {
            "role": "user",
            "content": [
                {"type": "text", "text": EXTRACTION_PROMPT},
                {
                    "type": "image_url",
                    "image_url": {"url": f"data:{media_type};base64,{b64}"},
                },
            ],
        }
    ]

    base_url = config["base_url"].rstrip("/")
    endpoint = f"{base_url}/chat/completions" if not base_url.endswith("/chat/completions") else base_url

    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            endpoint,
            headers={
                "Authorization": f"Bearer {config['api_key']}",
                "Content-Type": "application/json",
            },
            json={"model": config["model"], "messages": messages, "max_tokens": 4096},
        )

    if resp.status_code != 200:
        raise Exception(f"OpenAI {resp.status_code}: {resp.text[:200]}")

    data = resp.json()
    text_content = data.get("choices", [{}])[0].get("message", {}).get("content", "")
    return _parse_ai_response(text_content, filename, source, config)


def _parse_ai_response(text: str, filename: str, source: str, config: dict) -> dict | None:
    """Parse AI response text into structured data.

    Supports two response formats:
    - New: {"supplier": {...}, "document": {...}, "items": [...]}
    - Legacy: plain JSON array of items [...]
    """
    text = text.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)

    parsed = None
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        # Try to find JSON object or array in the response
        obj_match = re.search(r"\{.*\}", text, re.DOTALL)
        arr_match = re.search(r"\[.*\]", text, re.DOTALL)
        for m in [obj_match, arr_match]:
            if m:
                try:
                    parsed = json.loads(m.group())
                    break
                except json.JSONDecodeError:
                    continue

    if parsed is None:
        return None

    # Handle new format: {supplier, document, items}
    supplier_info = None
    document_info = None
    if isinstance(parsed, dict) and "items" in parsed:
        supplier_info = parsed.get("supplier")
        document_info = parsed.get("document")
        lines = parsed.get("items", [])
    elif isinstance(parsed, list):
        lines = parsed
    else:
        return None

    if not isinstance(lines, list) or not lines:
        return None

    result = {
        "lines": lines,
        "metadata": {
            "source": source,
            "filename": filename,
            "ai_provider": config.get("provider", "unknown"),
            "ai_model": config.get("model", "unknown"),
            "items_extracted": len(lines),
        },
    }
    if supplier_info and isinstance(supplier_info, dict) and supplier_info.get("name"):
        result["supplier"] = supplier_info
    if document_info and isinstance(document_info, dict):
        result["document"] = document_info

    return result
