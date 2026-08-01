"""Tests del exportador al Excel canonico de `ssa_construction_apu`.

El contrato con Odoo es el formato del archivo, asi que los tests generan el
libro, lo vuelven a abrir con openpyxl y verifican lo que el wizard va a leer:
hojas, cabeceras, refs cruzadas y precision de los numeros.

Se usan dobles (clases simples con los atributos que consume el exportador) en
vez de modelos SQLAlchemy: la funcion no toca la base de datos y los tests
tampoco deben hacerlo.
"""
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.apu_export import (
    SHEET_HEADERS,
    SHEET_ORDER,
    SUB_REF_PREFIX,
    ApuExportError,
    build_workbook,
    default_filename,
)


# ── Dobles de los modelos ──────────────────────────────────────
class _TemplateLine:
    def __init__(self, code, name, type, value=0.0, formula=None,
                 is_total=False, sequence=10, id=None):
        self.id = id
        self.code = code
        self.name = name
        self.type = type
        self.value = value
        self.formula = formula
        self.is_total = is_total
        self.sequence = sequence


class _Template:
    def __init__(self, name, lines):
        self.name = name
        self.lines = lines


class _Rubro:
    def __init__(self, id, name, sequence=10):
        self.id = id
        self.name = name
        self.sequence = sequence


class _Line:
    def __init__(self, type, name, quantity, price_unit, uom="u", insumo_id=None,
                 linked_item_id=None, notes=None, sequence=10, id=None):
        self.id = id
        self.type = type
        self.name = name
        self.quantity = quantity
        self.price_unit = price_unit
        self.uom = uom
        self.insumo_id = insumo_id
        self.linked_item_id = linked_item_id
        self.notes = notes
        self.sequence = sequence


class _Computo:
    def __init__(self, name, pieces=1.0, length=1.0, width=1.0, height=1.0,
                 sequence=10, id=None):
        self.id = id
        self.name = name
        self.pieces = pieces
        self.length = length
        self.width = width
        self.height = height
        self.sequence = sequence


class _Item:
    def __init__(self, id, name, rubro_id=None, uom="m2", quantity=1.0,
                 is_complementary=False, reference_price=None, unit_price=0.0,
                 sequence=10, lines=None, computos=None):
        self.id = id
        self.name = name
        self.rubro_id = rubro_id
        self.uom = uom
        self.quantity = quantity
        self.is_complementary = is_complementary
        self.reference_price = reference_price
        self.unit_price = unit_price
        self.sequence = sequence
        self.lines = lines or []
        self.computos = computos or []


class _Project:
    def __init__(self, rubros=None, items=None, template=None, name="Obra demo",
                 code="OBRA-01", decimals_price=2, decimals_performance=3,
                 decimals_subtotal=2, decimals_total=2, decimals_qty=4):
        self.name = name
        self.code = code
        self.rubros = rubros or []
        self.items = items or []
        self.template = template
        self.decimals_price = decimals_price
        self.decimals_performance = decimals_performance
        self.decimals_subtotal = decimals_subtotal
        self.decimals_total = decimals_total
        self.decimals_qty = decimals_qty


# ── Presupuesto de ejemplo ─────────────────────────────────────
def _template() -> _Template:
    return _Template(
        "GAMLP 2026",
        [
            _TemplateLine("MAT", "Materiales", "sum_mat", sequence=10),
            _TemplateLine("MO", "Mano de obra", "sum_mo", sequence=20),
            _TemplateLine("EQ", "Equipo", "sum_eq", sequence=30),
            _TemplateLine("B1", "Cargas sociales", "percent", value=71.18,
                          formula="MO", sequence=40),
            _TemplateLine("T", "Precio unitario", "formula",
                          formula="MAT + MO + EQ + B1", is_total=True,
                          sequence=50),
        ],
    )


def _project() -> _Project:
    """Obra con 2 rubros, una partida complementaria y un sub-APU."""
    r1 = _Rubro(1, "Obras preliminares", sequence=10)
    r2 = _Rubro(2, "Obra gruesa", sequence=20)

    # Partida complementaria: se consume como recurso 'sub' desde otra.
    mortero = _Item(
        10, "Mortero 1:4", rubro_id=2, uom="m3", quantity=1.0,
        is_complementary=True, sequence=10,
        lines=[
            _Line("mat", "Cemento portland IP-30", 7.5, 57.60, uom="bls",
                  insumo_id=100, sequence=10),
            _Line("mat", "Arena fina", 1.05, 120.0, uom="m3", insumo_id=101,
                  sequence=20),
        ],
    )

    muro = _Item(
        11, "Muro de ladrillo 6H", rubro_id=2, uom="m2", quantity=125.5,
        sequence=20,
        lines=[
            _Line("mat", "Ladrillo 6 huecos", 27.0, 3.20, uom="pza",
                  insumo_id=102, sequence=10),
            # Mismo insumo y mismo precio que en 'mortero' -> misma ref INSUMOS
            _Line("mat", "Cemento portland IP-30", 0.125, 57.60, uom="bls",
                  insumo_id=100, sequence=20),
            _Line("mo", "Albanil", 0.8, 18.75, uom="hr", insumo_id=200,
                  sequence=30, notes="incluye ayudante"),
            _Line("eq", "Herramientas menores", 0.05, 100.0, uom="glb",
                  sequence=40),
            _Line("sub", "Mortero 1:4", 0.025, 0.0, uom="m3",
                  linked_item_id=10, sequence=50),
        ],
        computos=[
            _Computo("Fachada norte", pieces=1, length=12.5, width=1.0,
                     height=2.8, sequence=10),
            _Computo("Fachada sur", pieces=2, length=10.0, width=1.0,
                     height=2.8, sequence=20),
        ],
    )

    excavacion = _Item(
        12, "Replanteo y trazado", rubro_id=1, uom="m2", quantity=340.0,
        sequence=5, reference_price=None, unit_price=8.45, lines=[],
    )

    return _Project(rubros=[r1, r2], items=[excavacion, mortero, muro],
                    template=_template())


def _load(data: bytes):
    return load_workbook(BytesIO(data), data_only=True)


def _rows(ws) -> list[list]:
    """Filas de datos (sin cabecera) como listas."""
    return [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]


def _dict_rows(ws) -> list[dict]:
    headers = [c.value for c in ws[1]]
    return [dict(zip(headers, r)) for r in _rows(ws)]


@pytest.fixture(scope="module")
def wb():
    return _load(build_workbook(_project(), _template()))


# ── Estructura del libro ───────────────────────────────────────
def test_devuelve_bytes_de_un_xlsx():
    data = build_workbook(_project(), _template())
    assert isinstance(data, bytes)
    assert data[:2] == b"PK"  # un .xlsx es un zip


def test_existen_las_siete_hojas_en_orden(wb):
    assert wb.sheetnames == list(SHEET_ORDER)
    assert len(wb.sheetnames) == 7


@pytest.mark.parametrize("sheet", SHEET_ORDER)
def test_cabeceras_exactas(wb, sheet):
    """Odoo localiza las columnas por nombre: un renombre rompe la carga."""
    encontradas = tuple(c.value for c in wb[sheet][1])
    assert encontradas == SHEET_HEADERS[sheet]


# ── CONFIG ─────────────────────────────────────────────────────
def test_config_lleva_las_precisiones_y_la_plantilla(wb):
    config = {k: v for k, v in _rows(wb["CONFIG"])}
    assert config["decimals_price"] == 2
    assert config["decimals_performance"] == 3
    assert config["decimals_subtotal"] == 2
    assert config["decimals_total"] == 2
    assert config["decimals_qty"] == 4
    assert config["plantilla_calc"] == "GAMLP 2026"


def test_config_respeta_precisiones_no_estandar():
    proyecto = _project()
    proyecto.decimals_price = 4
    proyecto.decimals_performance = 5
    config = {k: v for k, v in _rows(_load(build_workbook(proyecto, _template()))["CONFIG"])}
    assert config["decimals_price"] == 4
    assert config["decimals_performance"] == 5


# ── PLANTILLA_CALC ─────────────────────────────────────────────
def test_plantilla_exporta_las_filas_en_orden_y_marca_el_precio_final(wb):
    filas = _dict_rows(wb["PLANTILLA_CALC"])
    assert [f["codigo"] for f in filas] == ["MAT", "MO", "EQ", "B1", "T"]
    assert [f["seq"] for f in filas] == [10, 20, 30, 40, 50]

    finales = [f for f in filas if f["es_precio_final"]]
    assert len(finales) == 1
    assert finales[0]["codigo"] == "T"

    b1 = next(f for f in filas if f["codigo"] == "B1")
    assert b1["tipo"] == "percent"
    assert b1["valor_pct"] == pytest.approx(71.18)
    assert b1["formula"] == "MO"


def test_plantilla_solo_admite_los_tipos_de_odoo():
    tpl = _Template("Rota", [_TemplateLine("X", "X", "sum_todo")])
    with pytest.raises(ApuExportError):
        build_workbook(_project(), tpl)


def test_sin_plantilla_la_hoja_queda_solo_con_cabecera():
    proyecto = _project()
    proyecto.template = None
    hoja = _load(build_workbook(proyecto, None))["PLANTILLA_CALC"]
    assert tuple(c.value for c in hoja[1]) == SHEET_HEADERS["PLANTILLA_CALC"]
    assert _rows(hoja) == []


# ── RUBROS e ITEMS ─────────────────────────────────────────────
def test_rubros_usan_refs_legibles(wb):
    filas = _dict_rows(wb["RUBROS"])
    assert [f["ref"] for f in filas] == ["R1", "R2"]
    assert [f["nombre"] for f in filas] == ["Obras preliminares", "Obra gruesa"]
    assert [f["secuencia"] for f in filas] == [10, 20]


def test_items_apuntan_a_un_rubro_existente(wb):
    rubro_refs = {f["ref"] for f in _dict_rows(wb["RUBROS"])}
    for item in _dict_rows(wb["ITEMS"]):
        assert item["rubro_ref"] in rubro_refs


def test_items_salen_ordenados_por_secuencia(wb):
    filas = _dict_rows(wb["ITEMS"])
    assert [f["nombre"] for f in filas] == [
        "Replanteo y trazado", "Mortero 1:4", "Muro de ladrillo 6H",
    ]
    assert [f["ref"] for f in filas] == ["I1", "I2", "I3"]


def test_item_complementario_se_marca(wb):
    filas = {f["nombre"]: f for f in _dict_rows(wb["ITEMS"])}
    assert filas["Mortero 1:4"]["es_complementario"] == 1
    assert filas["Muro de ladrillo 6H"]["es_complementario"] == 0


def test_item_sin_lineas_conserva_su_precio_como_referencia(wb):
    """Sin composicion, el numero solo puede viajar en precio_referencia."""
    fila = next(f for f in _dict_rows(wb["ITEMS"]) if f["nombre"] == "Replanteo y trazado")
    assert fila["precio_referencia"] == pytest.approx(8.45)
    assert fila["estado"] == "draft"


# ── INSUMOS ────────────────────────────────────────────────────
def test_insumos_no_incluye_lineas_sub(wb):
    """Odoo autogenera el insumo del sub-APU desde la partida complementaria."""
    tipos = {f["tipo"] for f in _dict_rows(wb["INSUMOS"])}
    assert tipos <= {"mat", "mo", "eq"}
    assert "sub" not in tipos
    refs = {f["ref"] for f in _dict_rows(wb["INSUMOS"])}
    assert not any(r.startswith(SUB_REF_PREFIX) for r in refs)


def test_mismo_insumo_y_precio_se_agrupa_en_una_sola_ref(wb):
    """El cemento aparece en dos partidas al mismo precio: un solo recurso."""
    filas = _dict_rows(wb["INSUMOS"])
    cementos = [f for f in filas if f["nombre"] == "Cemento portland IP-30"]
    assert len(cementos) == 1

    lineas = _dict_rows(wb["ITEM_LINEAS"])
    usos = [l for l in lineas if l["insumo_ref"] == cementos[0]["ref"]]
    assert len(usos) == 2  # mortero y muro comparten la ref


def test_mismo_insumo_a_distinto_precio_son_recursos_distintos():
    """Dos proveedores, dos precios: colapsarlos falsearia el APU."""
    item = _Item(1, "Partida", lines=[
        _Line("mat", "Cemento", 1.0, 57.60, insumo_id=100, sequence=10),
        _Line("mat", "Cemento", 1.0, 62.00, insumo_id=100, sequence=20),
    ])
    hoja = _load(build_workbook(_Project(items=[item]), None))["INSUMOS"]
    filas = _dict_rows(hoja)
    assert len(filas) == 2
    assert {f["precio_unitario"] for f in filas} == {57.60, 62.00}
    assert len({f["ref"] for f in filas}) == 2


def test_insumos_tienen_ref_unica(wb):
    refs = [f["ref"] for f in _dict_rows(wb["INSUMOS"])]
    assert len(refs) == len(set(refs))
    assert all(r.startswith("INS") for r in refs)


# ── ITEM_LINEAS: integridad referencial ────────────────────────
def test_toda_linea_apunta_a_un_item_existente(wb):
    item_refs = {f["ref"] for f in _dict_rows(wb["ITEMS"])}
    for linea in _dict_rows(wb["ITEM_LINEAS"]):
        assert linea["item_ref"] in item_refs


def test_toda_linea_resuelve_su_insumo_o_sigue_la_convencion_sub(wb):
    """El requisito central: Odoo resuelve por texto, nada puede quedar huerfano."""
    insumo_refs = {f["ref"] for f in _dict_rows(wb["INSUMOS"])}
    item_names = {f["nombre"] for f in _dict_rows(wb["ITEMS"])}

    lineas = _dict_rows(wb["ITEM_LINEAS"])
    assert lineas  # el presupuesto de prueba tiene lineas

    for linea in lineas:
        ref = linea["insumo_ref"]
        if ref.startswith(SUB_REF_PREFIX):
            # La clave sintetica debe nombrar una partida real del libro.
            assert ref[len(SUB_REF_PREFIX):] in item_names
        else:
            assert ref in insumo_refs


def test_la_linea_sub_usa_el_nombre_de_la_partida_complementaria(wb):
    subs = [l for l in _dict_rows(wb["ITEM_LINEAS"])
            if l["insumo_ref"].startswith(SUB_REF_PREFIX)]
    assert len(subs) == 1
    assert subs[0]["insumo_ref"] == "__sub__Mortero 1:4"


def test_notas_de_linea_viajan(wb):
    lineas = _dict_rows(wb["ITEM_LINEAS"])
    assert any(l["notas"] == "incluye ayudante" for l in lineas)
    # Sin notas la celda queda en blanco (openpyxl la lee como None), nunca
    # con un "None" literal que Odoo importaria como texto.
    assert all(l["notas"] in (None, "") or isinstance(l["notas"], str)
               for l in lineas)
    assert "None" not in {l["notas"] for l in lineas}


def test_tipo_de_recurso_desconocido_se_rechaza():
    item = _Item(1, "Partida", lines=[_Line("otro", "X", 1.0, 1.0)])
    with pytest.raises(ApuExportError):
        build_workbook(_Project(items=[item]), None)


# ── COMPUTOS ───────────────────────────────────────────────────
def test_computos_apuntan_a_su_item(wb):
    item_refs = {f["ref"] for f in _dict_rows(wb["ITEMS"])}
    filas = _dict_rows(wb["COMPUTOS"])
    assert len(filas) == 2
    for fila in filas:
        assert fila["item_ref"] in item_refs
    assert [f["sector"] for f in filas] == ["Fachada norte", "Fachada sur"]


def test_computos_llevan_las_cuatro_dimensiones(wb):
    fila = next(f for f in _dict_rows(wb["COMPUTOS"]) if f["sector"] == "Fachada sur")
    assert fila["piezas"] == pytest.approx(2.0)
    assert fila["largo"] == pytest.approx(10.0)
    assert fila["ancho"] == pytest.approx(1.0)
    assert fila["alto"] == pytest.approx(2.8)


# ── Precision numerica ─────────────────────────────────────────
def test_el_rendimiento_sale_con_los_decimales_de_la_obra():
    """0.125 bolsas de cemento por m2 no puede truncarse a 0.13."""
    item = _Item(1, "Partida", lines=[
        _Line("mat", "Cemento", 0.1254, 10.0, insumo_id=1, sequence=10),
    ])
    proyecto = _Project(items=[item], decimals_performance=3)
    linea = _dict_rows(_load(build_workbook(proyecto, None))["ITEM_LINEAS"])[0]
    assert linea["rendimiento"] == 0.125


def test_el_precio_se_redondea_medio_arriba_como_odoo():
    """round() de Python daria 10.55 y el ERP mostraria 10.56."""
    item = _Item(1, "Partida", lines=[
        _Line("mat", "X", 1.0, 10.555, insumo_id=1),
    ])
    fila = _dict_rows(_load(build_workbook(_Project(items=[item]), None))["INSUMOS"])[0]
    assert fila["precio_unitario"] == 10.56


def test_la_cantidad_de_contrato_usa_decimals_qty():
    item = _Item(1, "Partida", quantity=125.123456)
    proyecto = _Project(items=[item], decimals_qty=4)
    fila = _dict_rows(_load(build_workbook(proyecto, None))["ITEMS"])[0]
    assert fila["cantidad_contrato"] == 125.1235


def test_las_precisiones_de_la_obra_mandan_sobre_el_default():
    item = _Item(1, "Partida", quantity=1.0, lines=[
        _Line("mat", "X", 0.123456, 10.987654, insumo_id=1),
    ])
    proyecto = _Project(items=[item], decimals_price=3, decimals_performance=2)
    libro = _load(build_workbook(proyecto, None))
    assert _dict_rows(libro["INSUMOS"])[0]["precio_unitario"] == 10.988
    assert _dict_rows(libro["ITEM_LINEAS"])[0]["rendimiento"] == 0.12


def test_las_celdas_numericas_declaran_su_formato():
    """El Excel viaja a una oficina tecnica: debe verse con sus decimales."""
    data = build_workbook(_project(), _template())
    libro = load_workbook(BytesIO(data))
    assert libro["ITEM_LINEAS"]["C2"].number_format == "0.000"
    assert libro["INSUMOS"]["E2"].number_format == "0.00"
    assert libro["COMPUTOS"]["C2"].number_format == "0.0000"


def test_los_numeros_van_como_numeros_no_como_texto(wb):
    for fila in _dict_rows(wb["INSUMOS"]):
        assert isinstance(fila["precio_unitario"], (int, float))
    for fila in _dict_rows(wb["ITEM_LINEAS"]):
        assert isinstance(fila["rendimiento"], (int, float))


# ── Casos borde ────────────────────────────────────────────────
def test_presupuesto_vacio_genera_un_libro_valido():
    libro = _load(build_workbook(_Project(), None))
    assert libro.sheetnames == list(SHEET_ORDER)
    for sheet in ("RUBROS", "INSUMOS", "ITEMS", "ITEM_LINEAS", "COMPUTOS"):
        assert _rows(libro[sheet]) == []


def test_item_sin_rubro_deja_la_ref_vacia():
    item = _Item(1, "Suelta", rubro_id=None)
    fila = _dict_rows(_load(build_workbook(_Project(items=[item]), None))["ITEMS"])[0]
    # Celda en blanco: openpyxl la devuelve como None, Odoo la lee como vacia.
    assert not fila["rubro_ref"]


def test_sin_proyecto_falla_explicito():
    with pytest.raises(ApuExportError):
        build_workbook(None, None)


def test_la_plantilla_se_toma_del_proyecto_si_no_se_pasa():
    proyecto = _project()
    config = {k: v for k, v in _rows(_load(build_workbook(proyecto))["CONFIG"])}
    assert config["plantilla_calc"] == "GAMLP 2026"


def test_nombre_de_archivo_sugerido():
    assert default_filename(_project()) == "APU_OBRA-01.xlsx"
    assert default_filename(_Project(code=None, name="Casa / Lote 3")).endswith(".xlsx")


def test_la_exportacion_es_determinista():
    """Dos corridas del mismo presupuesto deben dar las mismas refs."""
    a = _dict_rows(_load(build_workbook(_project(), _template()))["INSUMOS"])
    b = _dict_rows(_load(build_workbook(_project(), _template()))["INSUMOS"])
    assert a == b
